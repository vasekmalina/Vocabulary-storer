from bs4 import BeautifulSoup
import requests
import pyautogui as pya
import pyperclip 
import keyboard
from openpyxl import Workbook, load_workbook
import time
import string 
from rich.progress import track
import pathlib
path = str(pathlib.Path(__file__).parent.resolve())

#variables
e_words = []
c_words = []
letters = string.ascii_lowercase


def copy_clipboard():
    # this func copies clipboard
    pya.hotkey('ctrl', 'c')
    time.sleep(.01)
    return pyperclip.paste()

def save_words():
    # this func record words and store them in list
    words = []
    run = True
    keypress = False
    clipboard = ""
    while run:
        #executing when key is RELEASED
        if keypress and not keyboard.is_pressed("s"):
            clipboard = copy_clipboard()
            cb = ""
            for char in clipboard:
                if char in letters:
                    cb += char

            words.append(cb)
            print(clipboard)
            keypress = False
        elif keyboard.is_pressed("s") and not keypress:
            keypress = True

        elif keyboard.is_pressed("w"):
            time.sleep(0.5)
            put = input("Enter the word: ")
            words.append(put)
            print(put)
        elif keyboard.is_pressed('e'):
            run = False

    print(words)
    return words

def check_words():
    wb = load_workbook(path+"/ENG_vocab.xlsx")
    ws = wb["Input"]

    input_content = []
    words = []
    remove = []
    #check for duplicates in e_words
    for x in e_words:
        if x not in words:
            words.append(x)

    #check for duplicates in input sheet
    for row in ws.iter_rows():
        input_content.append(row[0].value)

    for w in words:
        for i in input_content:
            if w == i:
                remove.append(w)
    for r in remove:
        words.remove(r)
    
    return words



def translate():
    # this func takes eng words and translate them via Lingea dictionary using request and beautifulsoup library
    link = "https://slovniky.lingea.cz/anglicko-cesky/"
    for word in track(e_words, description= "Translating… "):
        page = link + word
        source = requests.get(page).text
        soup = BeautifulSoup(source, "lxml")
        c_trans = ""

        for article in soup.find_all("span", class_ = "lex_ful_tran w l2"):
            c_trans += str(article.text) + ", "

        if c_trans == "":
            c_trans = "Lingea toto slovo nezná."
        c_words.append(c_trans)

    return c_words



def save_in_sheet():
    # This func stores eng words as well as czech trans in excel sheet
    wb = load_workbook(path+"/ENG_vocab.xlsx")
    ws = wb["Input"]
    row = ws.max_row

    for e, c in zip(e_words, c_words):
        row += 1
        ws["A" + str(row)].value = e
        ws["B" + str(row)].value = c

    wb.save(path+"/ENG_vocab.xlsx")



#executing
e_words = save_words()
e_words = check_words()
c_words = translate()
save_in_sheet()
time.sleep(2)
