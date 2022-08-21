import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Color
import time
import pathlib
path = str(pathlib.Path(__file__).parent.resolve())


wb = load_workbook(path+"/ENG_vocab.xlsx")

#variables
input = []
new_words = []
unknown = []

#sheets
ws_input = wb["Input"]
ws_main = wb["Main"]
ws_unknown = wb["Unknown"]

try:
    #puting new word in list
    for row in ws_input.iter_rows():
        input.append([row[0].value, row[1].value, "*"])

    ws_input.delete_rows(1, ws_input.max_row)

    #check, whether there are same words in input
    for word in input:
        if word not in new_words:
            new_words.append(word)


    #deleting words that are already in main worksheet
    for row in ws_main["A"]:
        e_word = row.value
        for index, n_word in enumerate(new_words):
            n = n_word[0]
            if n == e_word:
                del new_words[index]


    #adding new words to main 
    for i in new_words:
        ws_main.append(i)

    #changing color to red by word that have "x" in 4. column
    font_style = Font(color= "D52904")
    for index, row in enumerate(ws_main["D"]):
        row_value = row.value
        if row_value == "x":
            var = ws_main["A" + str(index + 1)]
            var.font = font_style

    #uknown sheet prepar
    for row in ws_main.iter_rows():
        c = row[2].value
        d = row[3].value
        if len(c) > 2 and d != "x":
            unknown.append([row[0].value, row[1].value, row[2].value])


    #writing in unknown sheet
    ws_unknown.delete_rows(1, ws_unknown.max_row)
    for i in unknown:
        ws_unknown.append(i)

    print("Successfully completed.")
    time.sleep(3)
except:
    print("There has been an error.")
    time.sleep(5)

wb.save(path+"/ENG_vocab.xlsx")



