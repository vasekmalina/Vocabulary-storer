import openpyxl
from openpyxl import Workbook, load_workbook 
import math
import colorama
import collections

wb = load_workbook("C:/Users/vasek/Plocha/BIGY/vocabulary/ENG_vocab.xlsx")

ws = wb["Main"]
max = ws.max_row 

words = []
for row in ws.iter_rows():
    words.append(row[0].value)

if len(set(words)) == len(words):
    print("There are no duplicate words in your sheet.")
else:
    print("There are duplicate words in your sheet.")
    repeat = [item for item, count in collections.Counter(words).items() if count > 1]
    print("The words are:")
    for word in repeat:
        print(word, end=", ")

input()